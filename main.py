from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

siteUrl = 'https://leetcode.com/problem-list/top-interview-questions/'
questionNameList = []
questionUrlList = []
questionDifficultyList = []


def xcelSheet():

    df = pd.DataFrame({
        'Question Name': questionNameList,
        'Question Url': questionUrlList,
        'Question Difficulty': questionDifficultyList
    })
    wordfilename = 'Leetcode2.txt'
    
    with open(wordfilename,'w') as f:
        for i in range(0,df.__len__()):
            f.write(df['Question Name'][i] + ': ' + df['Question Url'][i] + ' Difficulty: ' + df['Question Difficulty'][i] + '\n')
    
    # wb = Workbook()
    # sheet1 = wb.create_sheet(sheetName)
    # sheet1.cell(1, 1, 'Question Name')
    # sheet1.cell(1, 2, 'Question URL')
    # sheet1.cell(1, 3, 'Question Difficulty')

    # for i in range(0, df.__len__()):
    #     sheet1.cell(i + 2, 1, df['Question Name'][i])
    #     sheet1.cell(i + 2, 2, df['Question Url'][i])
    #     sheet1.cell(i + 2, 3, df['Question Difficulty'][i])

    # wb.save(excelFileName)
    # wb.close()
    print("     -----------> word file created")


def openBrowser(url):
    print("     -----------> Opening Browser")
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_argument('--incognito')
    options.add_argument('--headless')

    # driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

    # headless browser
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)

    driver.get(url)
    driver.maximize_window()
    return driver


def closeBrowser(driver):
    print("     -----------> Closing Browser")
    driver.close()


# def fetchPageData(pageUrl):
#     sleepTime = 3

#     # print("Page URL: ", pageUrl)
#     browser = openBrowser(pageUrl)
#     time.sleep(sleepTime)
#     pageSource = browser.page_source
#     WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[role="rowgroup"]')))
#     WebDriverWait(browser, 10).until(EC.title_contains("Top Interview Questions - LeetCode"))
#     # print(f"title is: {browser.title}")

#     soup = BeautifulSoup(pageSource, 'html.parser')
#     print(
#             "\n\n                     ------------------- Parsing data -------------------\n\n"
#     )
#     newSoup = BeautifulSoup(pageSource, 'html.parser')
#     questionBlock = newSoup.find('div', role='rowgroup')
#     questionList = questionBlock.find_all('div', role='row')
#         # print(f"Total {questionList.__len__()} data fetched ")

#     for question in questionList:
#         row = question.find_all('div', role='cell')
#         questionName = row[1].find('a').text
#         questionUrl = row[1].find('a')['href']
#         questionUrl = 'https://leetcode.com' + questionUrl
#         questionDifficulty = row[4].find('span').text
#         questionNameList.append(questionName)
#         questionUrlList.append(questionUrl)
#         questionDifficultyList.append(questionDifficulty)
#             # print(questionName, questionUrl, questionDifficulty)
#     print("     -----------> Done")
#     closeBrowser(browser)

#     return

def fetchPageData(pageUrl):
    sleepTime = 3

    browser = openBrowser(pageUrl)
    time.sleep(sleepTime)

    WebDriverWait(browser, 10).until(EC.title_contains("Top Interview Questions - LeetCode"))
    print(f"\n\n------------------- Parsing data from {pageUrl} -------------------\n\n")

    # Wait for the question block to be present on the page
    WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[role="rowgroup"]')))

    pageSource = browser.page_source
    soup = BeautifulSoup(pageSource, 'html.parser')

    questionBlock = soup.find('div', role='rowgroup')
    questionList = questionBlock.find_all('div', role='row')

    for question in questionList:
        row = question.find_all('div', role='cell')
        questionName = row[1].find('a').text
        questionUrl = row[1].find('a')['href']
        questionUrl = 'https://leetcode.com' + questionUrl
        questionDifficulty = row[4].find('span').text
        questionNameList.append(questionName)
        questionUrlList.append(questionUrl)
        questionDifficultyList.append(questionDifficulty)

    print("-----------> Done")
    return browser

def getData():

    
        browser = openBrowser(siteUrl)
        time.sleep(2)
        pageSource = browser.page_source
        print(f"title is: {browser.title}")
        # print(f"page source:\n{pageSource}")

        # WebDriverWait(browser, 10).until(EC.title_contains("Top Interview Questions - LeetCode"))
        soup = BeautifulSoup(pageSource, 'html.parser')

            # Fetching total number of pages
        # soup = BeautifulSoup(html_content, 'html.parser')

# Find the nav component with the specified class
        nav_component = soup.find('nav', class_='mb-6 md:mb-0 flex flex-nowrap items-center space-x-2')
        # print(nav_component)
# Find all the buttons within the nav component
        buttons = nav_component.find_all('button')
    
# Extract the text or other attributes from each button
        for button in buttons:
            button_text = button.text
            print(button_text)
        # totalQuestion = soup.find('div', class_="text-label-2 dark:text-dark-label-2 mr-2").find_all('span')[1]
        # totalQuestion = totalQuestion.text.split('/')[1]
        # totalQuestion = int(totalQuestion)
            # print(f"Total {totalQuestion} questions available")
        totalPage = buttons[buttons.__len__() - 2].text
        totalPage=int(totalPage)
        print(f"Total {totalPage} pages available")
        closeBrowser(browser)

            # Fetching data from each page
        for page in range(1, totalPage + 1):
            print(
                    f"\n\n                     ------------------- Fetching Page {page} -------------------\n\n"
            )
            pageUrl = siteUrl + '?page=' + str(page)
            br=fetchPageData(pageUrl)
        closeBrowser(br)
        print("     -----------> Done all pages ")
        print(f"Total {questionNameList.__len__()} questions fetched")
        xcelSheet()

    


if __name__ == "__main__":
    getData() 